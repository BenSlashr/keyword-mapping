"""Service d'export pour générer les fichiers de résultats"""

import os
import logging
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..config import Config
from ..models import JobResult, Assignment, Keyword, CannibalAlert

logger = logging.getLogger(__name__)


class ExportService:
    """Service pour exporter les résultats vers différents formats"""
    
    def __init__(self):
        # Créer le dossier de résultats s'il n'existe pas
        os.makedirs(Config.RESULTS_DIR, exist_ok=True)
    
    async def export_to_xlsx(self, result: JobResult, job_id: str) -> str:
        """Exporte les résultats vers un fichier Excel formaté"""
        try:
            filename = f"keyword_matching_{job_id}.xlsx"
            filepath = os.path.join(Config.RESULTS_DIR, filename)
            
            # Créer un workbook avec plusieurs onglets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Onglet Summary
                summary_df = self._create_summary_dataframe(result)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Onglet Assignments
                assignments_df = self._create_assignments_dataframe(result.assignments)
                assignments_df.to_excel(writer, sheet_name='Assignments', index=False)
                
                # Onglet Orphans
                orphans_df = self._create_orphans_dataframe(result.orphans)
                orphans_df.to_excel(writer, sheet_name='Orphans', index=False)
                
                # Onglet Cannibalization
                cannibals_df = self._create_cannibals_dataframe(result.cannibals)
                cannibals_df.to_excel(writer, sheet_name='Cannibalization', index=False)
            
            # Appliquer le formatage Excel
            self._format_excel_file(filepath)
            
            logger.info(f"Export Excel créé: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur export Excel: {e}")
            raise
    
    async def export_to_csv(self, result: JobResult, job_id: str) -> str:
        """Exporte les résultats vers un fichier CSV"""
        try:
            filename = f"keyword_matching_{job_id}.csv"
            filepath = os.path.join(Config.RESULTS_DIR, filename)
            
            # Combiner toutes les données en un seul DataFrame
            all_data = []
            
            # Assignations
            for assignment in result.assignments:
                all_data.append({
                    'keyword': assignment.keyword,
                    'url': assignment.url,
                    'score': assignment.score,
                    'chunk_position': assignment.chunk_position,
                    'alternative_urls': '|'.join(assignment.alternative_urls),
                    'is_manual': assignment.is_manual,
                    'type': 'assigned'
                })
            
            # Orphelins
            for orphan in result.orphans:
                all_data.append({
                    'keyword': orphan.keyword,
                    'url': '',
                    'score': 0,
                    'chunk_position': '',
                    'alternative_urls': '',
                    'is_manual': False,
                    'type': 'orphan',
                    'volume': orphan.volume
                })
            
            # Cannibales
            for cannibal in result.cannibals:
                all_data.append({
                    'keyword': cannibal.keyword,
                    'url': cannibal.assigned_url,
                    'score': 0,
                    'chunk_position': '',
                    'alternative_urls': '',
                    'is_manual': False,
                    'type': 'cannibal',
                    'gsc_top_url': cannibal.gsc_top_url,
                    'gsc_clicks': cannibal.gsc_clicks,
                    'confidence_loss': cannibal.confidence_loss
                })
            
            df = pd.DataFrame(all_data)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            
            logger.info(f"Export CSV créé: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur export CSV: {e}")
            raise
    
    def _create_summary_dataframe(self, result: JobResult) -> pd.DataFrame:
        """Crée le DataFrame de résumé"""
        summary_data = [
            ['Metric', 'Value', 'Description'],
            ['Total Keywords', result.stats.get('total_keywords', 0), 'Nombre total de mots-clés traités'],
            ['Assigned Keywords', result.stats.get('assigned_keywords', 0), 'Mots-clés assignés avec succès'],
            ['Orphan Keywords', result.stats.get('orphan_keywords', 0), 'Mots-clés sans assignation'],
            ['Total Pages', result.stats.get('total_pages', 0), 'Nombre total de pages analysées'],
            ['Processing Time', f"{result.stats.get('processing_time_seconds', 0):.2f}s", 'Temps de traitement total'],
            ['Average Score', f"{result.stats.get('average_score', 0):.3f}", 'Score moyen des assignations'],
            ['Cannibalization Alerts', result.stats.get('cannibalization_alerts', 0), 'Alertes de cannibalisation détectées'],
            ['Assignment Rate', f"{(result.stats.get('assigned_keywords', 0) / max(result.stats.get('total_keywords', 1), 1) * 100):.1f}%", 'Taux d\'assignation'],
            ['Created At', result.created_at, 'Date de création du job'],
            ['Completed At', result.completed_at or 'N/A', 'Date de completion du job']
        ]
        
        return pd.DataFrame(summary_data[1:], columns=summary_data[0])
    
    def _create_assignments_dataframe(self, assignments: List[Assignment]) -> pd.DataFrame:
        """Crée le DataFrame des assignations"""
        data = []
        for assignment in assignments:
            data.append({
                'Keyword': assignment.keyword,
                'URL': assignment.url,
                'Score': round(assignment.score, 4),
                'Score (%)': f"{assignment.score * 100:.1f}%",
                'Chunk Position': assignment.chunk_position,
                'Alternative URL 1': assignment.alternative_urls[0] if len(assignment.alternative_urls) > 0 else '',
                'Alternative URL 2': assignment.alternative_urls[1] if len(assignment.alternative_urls) > 1 else '',
                'Alternative URL 3': assignment.alternative_urls[2] if len(assignment.alternative_urls) > 2 else '',
                'Is Manual': assignment.is_manual,
                'Confidence Level': self._get_confidence_level(assignment.score)
            })
        
        return pd.DataFrame(data)
    
    def _create_orphans_dataframe(self, orphans: List[Keyword]) -> pd.DataFrame:
        """Crée le DataFrame des mots-clés orphelins"""
        data = []
        for orphan in orphans:
            data.append({
                'Keyword': orphan.keyword,
                'Volume': orphan.volume if orphan.volume else 'N/A',
                'Reason': 'Score trop bas ou aucune page pertinente trouvée',
                'Suggestions': 'Créer du contenu spécifique ou optimiser les pages existantes'
            })
        
        return pd.DataFrame(data)
    
    def _create_cannibals_dataframe(self, cannibals: List[CannibalAlert]) -> pd.DataFrame:
        """Crée le DataFrame des alertes de cannibalisation"""
        data = []
        for cannibal in cannibals:
            data.append({
                'Keyword': cannibal.keyword,
                'Assigned URL': cannibal.assigned_url,
                'GSC Top URL': cannibal.gsc_top_url,
                'GSC Clicks': cannibal.gsc_clicks,
                'Confidence Loss': f"{cannibal.confidence_loss:.2f}",
                'Confidence Loss (%)': f"{cannibal.confidence_loss * 100:.1f}%",
                'Severity': self._get_cannibalization_severity(cannibal.confidence_loss),
                'Recommendation': self._get_cannibalization_recommendation(cannibal.confidence_loss)
            })
        
        return pd.DataFrame(data)
    
    def _get_confidence_level(self, score: float) -> str:
        """Détermine le niveau de confiance basé sur le score"""
        if score >= 0.8:
            return 'Très élevé'
        elif score >= 0.6:
            return 'Élevé'
        elif score >= 0.4:
            return 'Moyen'
        elif score >= 0.2:
            return 'Faible'
        else:
            return 'Très faible'
    
    def _get_cannibalization_severity(self, confidence_loss: float) -> str:
        """Détermine la sévérité de la cannibalisation"""
        if confidence_loss >= 0.7:
            return 'Critique'
        elif confidence_loss >= 0.5:
            return 'Élevée'
        elif confidence_loss >= 0.3:
            return 'Modérée'
        else:
            return 'Faible'
    
    def _get_cannibalization_recommendation(self, confidence_loss: float) -> str:
        """Donne une recommandation basée sur la sévérité"""
        if confidence_loss >= 0.7:
            return 'Action immédiate requise - Redirection ou consolidation de contenu'
        elif confidence_loss >= 0.5:
            return 'Optimiser le contenu de la page assignée ou rediriger'
        elif confidence_loss >= 0.3:
            return 'Surveiller et optimiser si nécessaire'
        else:
            return 'Impact minimal - Surveillance recommandée'
    
    def _format_excel_file(self, filepath: str):
        """Applique le formatage à un fichier Excel"""
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            # Styles de formatage
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Formater chaque feuille
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Formatage des en-têtes
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                
                # Formatage des données
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        
                        # Centrer certaines colonnes
                        if cell.column_letter in ['C', 'D', 'E', 'I', 'J']:
                            cell.alignment = center_alignment
                
                # Ajuster la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Figer la première ligne
                worksheet.freeze_panes = 'A2'
            
            # Sauvegarder
            workbook.save(filepath)
            logger.info(f"Formatage Excel appliqué: {filepath}")
            
        except Exception as e:
            logger.error(f"Erreur formatage Excel: {e}")
    
    async def export_to_json(self, result: JobResult, job_id: str) -> str:
        """Exporte les résultats vers un fichier JSON"""
        try:
            filename = f"keyword_matching_{job_id}.json"
            filepath = os.path.join(Config.RESULTS_DIR, filename)
            
            # Convertir le résultat en dictionnaire
            result_dict = result.dict()
            
            # Ajouter des métadonnées
            result_dict['export_metadata'] = {
                'exported_at': datetime.utcnow().isoformat(),
                'export_version': '2.0.0',
                'job_id': job_id
            }
            
            # Écrire le fichier JSON
            import json
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(result_dict, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Export JSON créé: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur export JSON: {e}")
            raise
    
    async def generate_html_report(self, result: JobResult, job_id: str) -> str:
        """Génère un rapport HTML interactif"""
        try:
            filename = f"keyword_matching_report_{job_id}.html"
            filepath = os.path.join(Config.RESULTS_DIR, filename)
            
            # Template HTML simple
            html_content = self._generate_html_template(result, job_id)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"Rapport HTML créé: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erreur génération rapport HTML: {e}")
            raise
    
    def _generate_html_template(self, result: JobResult, job_id: str) -> str:
        """Génère le template HTML pour le rapport"""
        
        # Statistiques de base
        stats = result.stats
        total_keywords = stats.get('total_keywords', 0)
        assigned_keywords = stats.get('assigned_keywords', 0)
        assignment_rate = (assigned_keywords / max(total_keywords, 1)) * 100
        
        html = f"""
        <!DOCTYPE html>
        <html lang="fr">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Rapport Keyword-URL Matcher - {job_id}</title>
            <script src="https://cdn.tailwindcss.com"></script>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
        </head>
        <body class="bg-gray-50 text-gray-900">
            <div class="container mx-auto px-4 py-8">
                <header class="mb-8">
                    <h1 class="text-3xl font-bold text-gray-800 mb-2">
                        Rapport d'Assignation Keyword-URL
                    </h1>
                    <p class="text-gray-600">Job ID: {job_id} | Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}</p>
                </header>
                
                <!-- Statistiques principales -->
                <div class="grid grid-cols-1 md:grid-cols-4 gap-6 mb-8">
                    <div class="bg-white rounded-lg shadow-md p-6">
                        <h3 class="text-lg font-semibold text-gray-700 mb-2">Mots-clés assignés</h3>
                        <p class="text-3xl font-bold text-blue-600">{assigned_keywords}</p>
                        <p class="text-sm text-gray-500">{assignment_rate:.1f}% du total</p>
                    </div>
                    <div class="bg-white rounded-lg shadow-md p-6">
                        <h3 class="text-lg font-semibold text-gray-700 mb-2">Pages analysées</h3>
                        <p class="text-3xl font-bold text-green-600">{stats.get('total_pages', 0)}</p>
                    </div>
                    <div class="bg-white rounded-lg shadow-md p-6">
                        <h3 class="text-lg font-semibold text-gray-700 mb-2">Mots-clés orphelins</h3>
                        <p class="text-3xl font-bold text-yellow-600">{stats.get('orphan_keywords', 0)}</p>
                    </div>
                    <div class="bg-white rounded-lg shadow-md p-6">
                        <h3 class="text-lg font-semibold text-gray-700 mb-2">Cannibalisations</h3>
                        <p class="text-3xl font-bold text-red-600">{stats.get('cannibalization_alerts', 0)}</p>
                    </div>
                </div>
                
                <!-- Graphique de distribution -->
                <div class="bg-white rounded-lg shadow-md p-6 mb-8">
                    <h3 class="text-xl font-semibold text-gray-800 mb-4">Distribution des Scores</h3>
                    <canvas id="scoreChart" width="400" height="200"></canvas>
                </div>
                
                <!-- Top assignations -->
                <div class="bg-white rounded-lg shadow-md p-6">
                    <h3 class="text-xl font-semibold text-gray-800 mb-4">Top 10 Assignations</h3>
                    <div class="overflow-x-auto">
                        <table class="min-w-full table-auto">
                            <thead class="bg-gray-50">
                                <tr>
                                    <th class="px-4 py-2 text-left">Mot-clé</th>
                                    <th class="px-4 py-2 text-left">URL</th>
                                    <th class="px-4 py-2 text-center">Score</th>
                                    <th class="px-4 py-2 text-center">Confiance</th>
                                </tr>
                            </thead>
                            <tbody>
        """
        
        # Top 10 assignations par score
        top_assignments = sorted(result.assignments, key=lambda x: x.score, reverse=True)[:10]
        
        for assignment in top_assignments:
            confidence = self._get_confidence_level(assignment.score)
            confidence_class = {
                'Très élevé': 'text-green-600',
                'Élevé': 'text-blue-600', 
                'Moyen': 'text-yellow-600',
                'Faible': 'text-orange-600',
                'Très faible': 'text-red-600'
            }.get(confidence, 'text-gray-600')
            
            html += f"""
                                <tr class="border-b">
                                    <td class="px-4 py-2 font-medium">{assignment.keyword}</td>
                                    <td class="px-4 py-2 text-blue-600">
                                        <a href="{assignment.url}" target="_blank" class="hover:underline">
                                            {assignment.url[:50]}{'...' if len(assignment.url) > 50 else ''}
                                        </a>
                                    </td>
                                    <td class="px-4 py-2 text-center font-mono">{assignment.score:.3f}</td>
                                    <td class="px-4 py-2 text-center">
                                        <span class="{confidence_class} font-semibold">{confidence}</span>
                                    </td>
                                </tr>
            """
        
        html += """
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            
            <script>
                // Graphique de distribution des scores
                const ctx = document.getElementById('scoreChart').getContext('2d');
                
                // Données pour le graphique (à remplacer par les vraies données)
                const scoreData = {
                    labels: ['0.0-0.2', '0.2-0.4', '0.4-0.6', '0.6-0.8', '0.8-1.0'],
                    datasets: [{
                        label: 'Nombre d\\'assignations',
                        data: [0, 0, 0, 0, 0], // À calculer depuis les vraies données
                        backgroundColor: [
                            'rgba(239, 68, 68, 0.8)',
                            'rgba(245, 158, 11, 0.8)',
                            'rgba(59, 130, 246, 0.8)',
                            'rgba(16, 185, 129, 0.8)',
                            'rgba(34, 197, 94, 0.8)'
                        ],
                        borderWidth: 1
                    }]
                };
                
                new Chart(ctx, {
                    type: 'bar',
                    data: scoreData,
                    options: {
                        responsive: true,
                        scales: {
                            y: {
                                beginAtZero: true
                            }
                        }
                    }
                });
            </script>
        </body>
        </html>
        """
        
        return html 